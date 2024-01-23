import os
import datetime
import requests
import time
import openpyxl
from openpyxl import Workbook

## Author: MrSh4d0w

i = 3
primeraejec = True
niteracion = 1
# ARCHIVO EXCEL

path = "...\\btcCMAPIPrice.xlsx" ## ESCRIBIR RUTA DEL ARCHIVO
pathexiste = os.path.exists(path)

if pathexiste == True:
    print('El archivo si existe asi que no hace falta crearlo')
    wb = openpyxl.load_workbook(path)

    hojaBTC = wb.active
    # Nombre hoja BTC (esto solo la primera vez)
    hojaBTC.title = "BTC"
    a1 = hojaBTC.cell(row=1, column=1)
    creado = False
else:
    # Creacion archivo
    print('No existe el archivo, asi que lo creamos')
    wb: Workbook = openpyxl.Workbook("btcCMAPIPrice.xlsx")
    #wb = openpyxl.load_workbook(path)
    wb.save('btcCMAPIPrice.xlsx')

    hojaBTC = wb.active
    # Nombre hoja BTC (esto solo la primera vez)
    hojaBTC.title = "BTC"
    creado = True

    #a1 = hojaBTC.cell(row=1, column=1, value=1)

# hojaBTC.row_dimensions[1].width = 20


def escribirexcelDate(fila, precio):
    hojaBTC.cell(row=fila, column=1, value=precio)


def escribirexcel(fila, precio):
    hojaBTC.cell(row=fila, column=1, value=precio).number_format = '0.00'

while True:

    headers = {
            'X-CMC_PRO_API_KEY': '', ##PONER AQUí API DE COINMARKETCAP
            'Accepts': 'application/json'
            }

    params = {
            'start': '1',
            'limit': '6',
            'convert': 'EUR'
            }

    url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'

    json = requests.get(url, params = params, headers = headers).json()

    coins = json['data']

    for coin in coins:
        if coin['symbol'] == 'BTC':
            btcprice: object = round(coin['quote']['EUR']['price'],4)

    a1 = hojaBTC.cell(row=1, column=1)

    if btcprice != None:
            a1 = hojaBTC.cell(row=1, column=1)

            # Si existe a1 lo usamos, y si no le damos el valor de 1, ya que implicará que no hay nada escrito
            if(a1.value == None):
                a1.value = 1
                row = a1.value
            else:
                row = a1.value


            if primeraejec:
                print('Primera ejecucion, último valor escrito en ' + str(row) + ' y empezará por ' + str(row+2))
                print('Fecha y hora ' + str(datetime.datetime.now()) + ' en fila ' + str(row+2))
                print("Iteracion " + str(niteracion) + ", fila " + str(row + i) + ", precio: " + str(btcprice))

                # Escribo la fila de la primera iteracion
                escribirexcel(1, row+3)

                # Escribo la fecha y el precio
                escribirexcelDate(row+2, datetime.datetime.now())
                #wb.number_format = '0.00'
                escribirexcel(row+3, btcprice)
                primeraejec = False

            else:
                i = a1.value+1
                print("Iteracion " + str(niteracion) + ", fila " + str(row+1) + ", precio: " + str(btcprice))
                escribirexcel(1, row+1)
                escribirexcel(row+1, float(btcprice))
    else:
        print("Error en el valor obtenido de BTC")

    row += 1
    niteracion += 1

    wb.save('btcCMAPIPrice.xlsx')
    wb.close()

    time.sleep(5)
